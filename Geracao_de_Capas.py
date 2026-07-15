import os
import re
import copy
from datetime import datetime
import openpyxl
import pandas as pd


def atualizar_datas_arquivos(nova_data_input=None):
    # 1. Solicita o caminho da pasta e a nova data ao usuário
    caminho_pasta = r"C:\Users\Hall\Documents\Python - HOMOLOGACAO\Capas Python"
    if nova_data_input is None:
        nova_data_input = input("Digite a nova data (DD/MM/AAAA): ").strip()

    if not os.path.exists(caminho_pasta):
        print("Erro: O caminho da pasta especificado não existe.")
        return
    nova_data = None
    try:
        nova_data = datetime.strptime(nova_data_input, "%d/%m/%Y").strftime(
            "%m-%Y"
        )
    except ValueError:
        pass

    if not nova_data:
        try:
            nova_data = datetime.strptime(nova_data_input, "%m/%Y").strftime(
                "%m-%Y"
            )
        except ValueError:
            pass
    if not nova_data:
        if re.match(r"^\d{2}-\d{4}$", nova_data_input):
            nova_data = nova_data_input
        else:
            print(
                "Erro: O formato de data digitado é inválido. Use DD/MM/AAAA ou MM/AAAA."
            )
            return
            
    print(f"\n-> Data formatada para o padrão dos arquivos: {nova_data}")
    
    contador_atualizados = 0
    print("\nIniciando a atualização dos arquivos...")
    print("-" * 50)
    
    for nome_arquivo in os.listdir(caminho_pasta):
        caminho_completo_antigo = os.path.join(caminho_pasta, nome_arquivo)
        
        if os.path.isfile(caminho_completo_antigo) and nome_arquivo.endswith(".xlsx"):
            # Expressão regular melhorada:
            # Procura qualquer variação de data no final (ex: 06-2026, 06.2026, 06/2026)
            padrao_data_existente = re.search(r"[-_\s]?\d{2}[-./]\d{4}$", os.path.splitext(nome_arquivo)[0])
            
            nome_base, extensao = os.path.splitext(nome_arquivo)
            
            if padrao_data_existente:
                # Se já tem uma data no final, substitui ela mantendo o resto do nome
                nome_sem_data = os.path.splitext(nome_arquivo)[0][:padrao_data_existente.start()]
                novo_nome_arquivo = f"{nome_sem_data} {nova_data}{extensao}"
            else:
                # Se o arquivo NÃO tem data nenhuma, apenas adiciona a nova data no final
                novo_nome_arquivo = f"{nome_base} {nova_data}{extensao}"
                
            caminho_completo_novo = os.path.join(caminho_pasta, novo_nome_arquivo)
            
            # Só renomeia se o nome for de fato diferente (evita erro se rodar o script duas vezes seguidas)
            if nome_arquivo != novo_nome_arquivo:
                os.rename(caminho_completo_antigo, caminho_completo_novo)
                print(f"Atualizado: '{nome_arquivo}' -> '{novo_nome_arquivo}'")
                contador_atualizados += 1
            else:
                # Conta como atualizado/validado se ele já estiver no formato correto
                contador_atualizados += 1

    print("-" * 50)
    print(
        f"Processo concluído! Total de {contador_atualizados} arquivos validados/atualizados."
    )
    return caminho_pasta, nova_data_input

def preencher_informacoes_planilha(caminho_pasta, nova_data_input, quem_concilia=None, quem_aprova=None, quem_reconcilia=None):
    print("\n--- Coleta de Informações para a Planilha ---")
    if quem_concilia is None:
        quem_concilia = input("Quem Concilia? ").strip()
    if quem_aprova is None:
        quem_aprova = input("Quem Aprova? ").strip()
    if quem_reconcilia is None:
        quem_reconcilia = input("Quem Reconcilia? ").strip()

    contador_planilhas = 0

    print("\nIniciando o preenchimento das planilhas...")
    print("-" * 50)

    # Varre a pasta procurando apenas arquivos Excel (.xlsx)
    for nome_arquivo in os.listdir(caminho_pasta):
        caminho_arquivo = os.path.join(caminho_pasta, nome_arquivo)

        # Verifica se é um arquivo e se possui a extensão .xlsx
        if os.path.isfile(caminho_arquivo) and nome_arquivo.endswith(".xlsx"):
            try:
                # Abre o arquivo Excel
                wb = openpyxl.load_workbook(caminho_arquivo)

                # Verifica se a aba "CONCILIAÇÃO" existe no arquivo
                if "CONCILIAÇÃO" in wb.sheetnames:
                    ws = wb["CONCILIAÇÃO"]

                    # Escreve as variáveis nas respectivas células

                    # --- TRATAMENTO DE DATA IDÊNTICO AO DO EPC ---
                    try:
                        data_objeto = pd.to_datetime(nova_data_input, dayfirst=True, errors='coerce')
                        if pd.notna(data_objeto):
                            # Injeta o objeto DATE nativo do Python
                            ws["B2"] = data_objeto.to_pydatetime().date()
                            ws["B2"].number_format = 'yyyy-mm-dd'  # Formato interno que o Excel traduz localmente
                        else:
                            ws["B2"] = nova_data_input
                    except Exception:
                        ws["B2"] = nova_data_input

                    ws["L1"] = quem_concilia
                    ws["L2"] = quem_aprova
                    ws["L3"] = quem_reconcilia

                    # Salva as alterações no arquivo
                    wb.save(caminho_arquivo)
                    print(
                        f"Informações gravadas com sucesso em: '{nome_arquivo}' (Aba: CONCILIAÇÃO)"
                    )
                    contador_planilhas += 1
                else:
                    print(
                        f"Aviso: Aba 'CONCILIAÇÃO' não encontrada no arquivo '{nome_arquivo}'."
                    )

                # Fecha o arquivo para liberar memória
                wb.close()

            except Exception as e:
                print(f"Erro ao processar o arquivo '{nome_arquivo}': {e}")

    print("-" * 50)
    print(
        f"Processo concluído! Total de {contador_planilhas} planilhas atualizadas."
    )

def importar_dados_balancete(caminho_pasta_capas, nova_data):
    # 1. Define o caminho da pasta do balancete (BOF)
    pasta_bof = r"C:\Users\Hall\Documents\Python - HOMOLOGACAO\Capas Python\Arquivos"

    print(f"\n--- Buscando Balancete (BOF) para a data {nova_data} ---")

    # 2. Localiza o arquivo aceitando tanto .xlsx quanto .csv
    arquivo_bof = None
    extensao_bof = None
    
    if os.path.exists(pasta_bof):
        for arquivo in os.listdir(pasta_bof):
            # Procura por "BOF" e a data "06-2026" no nome do arquivo
            if "BOF" in arquivo and nova_data in arquivo:
                if arquivo.endswith(".xlsx") or arquivo.endswith(".csv"):
                    arquivo_bof = os.path.join(pasta_bof, arquivo)
                    extensao_bof = os.path.splitext(arquivo)[1].lower()
                    break

    if not arquivo_bof:
        print(
            f"Erro: Arquivo do Balancete contendo 'BOF' e '{nova_data}' (.xlsx ou .csv) não foi encontrado em: {pasta_bof}"
        )
        return

    print(f"Balancete localizado: {os.path.basename(arquivo_bof)}")

    # 3. Lê o Balancete dinamicamente de acordo com a extensão do arquivo
    try:
        if extensao_bof == ".csv":
            # Lê arquivos .csv (Geralmente usa ponto e vírgula como separador no Brasil)
            df = pd.read_csv(arquivo_bof, sep=None, engine='python', encoding='latin-1')
        else:
            # Lê arquivos .xlsx
            df = pd.read_excel(arquivo_bof)
    except Exception as e:
        print(f"Erro ao ler o arquivo BOF: {e}")
        return

    # Validação de Colunas Necessárias
    colunas_obrigatorias = [
        "sCodContaContabil",
        "sNomeContaContabil",
        "nSaldoInicial",
        "nMovimDebito",
        "nMovimCredito",
        "nSaldoFinal",
    ]
    if not all(col in df.columns for col in colunas_obrigatorias):
        print(
            f"Erro: O balancete não possui todas as colunas necessárias. Colunas encontradas: {list(df.columns)}"
        )
        return

    print("Processando contas e atualizando capas de conciliação...")
    print("-" * 50)

    # Mapeia todos os arquivos .xlsx disponíveis na pasta principal de Capas para busca rápida
    arquivos_pasta_capas = os.listdir(caminho_pasta_capas)

    for _, linha in df.iterrows():
        conta_original = str(linha["sCodContaContabil"]).strip()

        conta_limpa = re.sub(r"[-.\s]", "", conta_original)

        conta_11_digitos = conta_limpa[:11]

        if not conta_11_digitos.isdigit():
            continue

        arquivo_capa_alvo = None
        for arq in arquivos_pasta_capas:
            if arq.startswith(conta_11_digitos) and arq.endswith(".xlsx"):
                arquivo_capa_alvo = os.path.join(caminho_pasta_capas, arq)
                break

        if arquivo_capa_alvo:
            try:
                wb = openpyxl.load_workbook(arquivo_capa_alvo)

                if "CONCILIAÇÃO" in wb.sheetnames:
                    ws = wb["CONCILIAÇÃO"]

                    ws["B1"] = linha["sNomeContaContabil"]

                    # --- TRATAMENTO NUMÉRICO IDÊNTICO AO DO EPC ---
                    for celula_ref, valor_bof in (
                        ("E1", linha["nSaldoInicial"]),
                        ("E2", linha["nMovimDebito"]),
                        ("E3", linha["nMovimCredito"]),
                        ("E4", linha["nSaldoFinal"]),
                    ):
                        celula = ws[celula_ref]

                        if pd.isna(valor_bof) or str(valor_bof).strip() == "" or str(valor_bof).lower() == "nan":
                            celula.value = None
                            continue

                        valor_str = str(valor_bof).strip()

                        try:
                            # Limpeza profunda de espaços e caracteres regionais
                            valor_str = valor_str.replace(" ", "")

                            # Trata se o sinal de menos estiver no fim (ex: 1.500,00-)
                            if valor_str.endswith("-"):
                                valor_str = "-" + valor_str[:-1]

                            if "," in valor_str:
                                valor_str = valor_str.replace(".", "").replace(",", ".")

                            # CONVERSÃO PARA FLOAT REAL (Garante que vira número para fórmulas)
                            celula.value = float(valor_str)
                            celula.number_format = '#,##0.00'
                        except Exception:
                            # Se falhar, tenta remover qualquer outro caractere residual
                            try:
                                limpo = re.sub(r'[^\d,.-]', '', valor_str)
                                if "," in limpo:
                                    limpo = limpo.replace(".", "").replace(",", ".")
                                celula.value = float(limpo)
                                celula.number_format = '#,##0.00'
                            except:
                                celula.value = valor_bof  # Fallback final

                    wb.save(arquivo_capa_alvo)
                    print(
                        f"Conta {conta_11_digitos}: Dados injetados com sucesso em '{os.path.basename(arquivo_capa_alvo)}'"
                    )
                else:
                    print(
                        f"Aviso: Aba 'CONCILIAÇÃO' não existe em '{os.path.basename(arquivo_capa_alvo)}'"
                    )

                wb.close()
            except Exception as e:
                print(
                    f"Erro ao atualizar a capa da conta {conta_11_digitos}: {e}"
                )

    print("-" * 50)
    print("Processamento do Balancete concluído!")

def importar_dados_epc(caminho_pasta_capas, nova_data):
    # 1. Define o caminho da pasta do relatório EPC
    pasta_epc = r"C:\Users\Hall\Documents\Python - HOMOLOGACAO\Capas Python\Arquivos"

    print(f"\n--- Buscando Relatório EPC para a data {nova_data} ---")

    # 2. Localiza o arquivo EPC (aceitando .xlsx ou .csv)
    arquivo_epc = None
    extensao_epc = None
    
    if os.path.exists(pasta_epc):
        for arquivo in os.listdir(pasta_epc):
            if "EPC" in arquivo.upper() and nova_data in arquivo:
                if arquivo.endswith(".xlsx") or arquivo.endswith(".csv"):
                    arquivo_epc = os.path.join(pasta_epc, arquivo)
                    extensao_epc = os.path.splitext(arquivo)[1].lower()
                    break

    if not arquivo_epc:
        print(f"Erro: Arquivo EPC contendo 'EPC' e '{nova_data}' (.xlsx ou .csv) não foi encontrado em: {pasta_epc}")
        return

    print(f"Relatório EPC localizado: {os.path.basename(arquivo_epc)}")

    # 3. Carrega os dados do EPC usando Pandas mantendo formato textual para tratamento limpo
    try:
        if extensao_epc == ".csv":
            df_epc = pd.read_csv(arquivo_epc, sep=None, engine='python', encoding='latin-1', dtype=str)
        else:
            df_epc = pd.read_excel(arquivo_epc, dtype=str)
    except Exception as e:
        print(f"Erro ao ler o arquivo EPC: {e}")
        return

    if df_epc.empty:
        print("Aviso: O arquivo EPC localizado está vazio.")
        return

    # Mapeamento exato das colunas que precisam de tratamento numérico e de data
    colunas_numero = ["nVlrInicio", "nDespApropriar", "nDespApropriada", "nVlrAtual", 
                      "nVlrContabilAtual", "nVlrComissaoCustoVendaApropriar", "nApropriacaoCustoVenda",
                      "nVlrResgate", "nDespDiaria"]
    colunas_data = ["dtDataBase", "dtDtVcto", "dtDtEmissao"]

    indices_numero = [df_epc.columns.get_loc(col) for col in colunas_numero if col in df_epc.columns]
    indices_data = [df_epc.columns.get_loc(col) for col in colunas_data if col in df_epc.columns]

    cabecalho = list(df_epc.columns)
    linhas_dados = df_epc.values.tolist()

    print("Copiando dados do EPC com tratamento numérico real para a aba 'EMISSÃO PROPRIA COMPLETO'...")
    print("-" * 50)

    # 4. Varre todas as capas da pasta principal para colar o relatório
    contador_capas = 0
    for nome_arquivo in os.listdir(caminho_pasta_capas):
        caminho_capa = os.path.join(caminho_pasta_capas, nome_arquivo)

        if os.path.isfile(caminho_capa) and nome_arquivo.endswith(".xlsx"):
            try:
                wb = openpyxl.load_workbook(caminho_capa)

                if "EMISSÃO PROPRIA COMPLETO" in wb.sheetnames:
                    ws = wb["EMISSÃO PROPRIA COMPLETO"]

                    # Limpa completamente dados e formatações antigas a partir da linha 30
                    for linha_limpar in range(30, ws.max_row + 1):
                        for col_limpar in range(1, len(cabecalho) + 1):
                            ws.cell(row=linha_limpar, column=col_limpar).value = None
                            ws.cell(row=linha_limpar, column=col_limpar).number_format = 'General'

                    # 1. Cola o Cabeçalho na Linha 30
                    for c_idx, valor in enumerate(cabecalho):
                        celula = ws.cell(row=30, column=c_idx + 1)
                        celula.value = valor

                    # 2. Cola as Linhas de Dados a partir da Linha 31
                    for r_idx, linha_dados in enumerate(linhas_dados):
                        linha_atual = 31 + r_idx
                        for c_idx, valor in enumerate(linha_dados):
                            celula = ws.cell(row=linha_atual, column=c_idx + 1)
                            
                            if pd.isna(valor) or str(valor).strip() == "" or str(valor).lower() == "nan":
                                celula.value = None
                                continue

                            valor_str = str(valor).strip()

                            # --- TRATAMENTO NUMÉRICO CRÍTICO ---
                            if c_idx in indices_numero:
                                try:
                                    # Limpeza profunda de espaços e caracteres regionais
                                    valor_str = valor_str.replace(" ", "")
                                    
                                    # Trata se o sinal de menos estiver no fim (ex: 1.500,00-)
                                    if valor_str.endswith("-"):
                                        valor_str = "-" + valor_str[:-1]
                                        
                                    if "," in valor_str:
                                        valor_str = valor_str.replace(".", "").replace(",", ".")
                                    
                                    # CONVERSÃO PARA FLOAT REAL (Garante que vira número para fórmulas)
                                    celula.value = float(valor_str)
                                    celula.number_format = '#,##0.00'
                                except Exception:
                                    # Se falhar, tenta remover qualquer outro caractere residual
                                    try:
                                        limpo = re.sub(r'[^\d,.-]', '', valor_str)
                                        if "," in limpo:
                                            limpo = limpo.replace(".", "").replace(",", ".")
                                        celula.value = float(limpo)
                                        celula.number_format = '#,##0.00'
                                    except:
                                        celula.value = valor  # Fallback final
                                
                            # --- TRATAMENTO DE DATAS CRÍTICO ---
                            elif c_idx in indices_data:
                                try:
                                    data_objeto = pd.to_datetime(valor_str, dayfirst=True, errors='coerce')
                                    if pd.notna(data_objeto):
                                        # Injeta o objeto DATE nativo do Python
                                        celula.value = data_objeto.to_pydatetime().date()
                                        celula.number_format = 'yyyy-mm-dd' # Formato interno que o Excel traduz localmente
                                    else:
                                        celula.value = valor
                                except Exception:
                                    celula.value = valor
                                
                            # Demais colunas de texto normal
                            else:
                                celula.value = valor_str

                    # --- CORREÇÃO DO ### (AUTO-AJUSTE DA LARGURA DAS COLUNAS) ---
                    for col in ws.columns:
                        max_len = 0
                        col_letter = openpyxl.utils.get_column_letter(col[0].column)
                        for cell in col:
                            if cell.row >= 30 and cell.value:
                                max_len = max(max_len, len(str(cell.value)))
                        if max_len > 0:
                            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

                    wb.save(caminho_capa)
                    print(f"Dados do EPC formatados e colados em: '{nome_arquivo}'")
                    contador_capas += 1

                wb.close()
            except Exception as e:
                print(f"Erro ao colar EPC no arquivo '{nome_arquivo}': {e}")

    print("-" * 50)
    print(f"Processamento do EPC concluído! Total de {contador_capas} capas atualizadas.")

def importar_dados_eps(caminho_pasta_capas, nova_data):
    # 1. Define o caminho da pasta do relatório EPS (mesma pasta do EPC)
    pasta_eps = r"C:\Users\Hall\Documents\Python - HOMOLOGACAO\Capas Python\Arquivos"

    print(f"\n--- Buscando Relatório EPS para a data {nova_data} ---")

    # 2. Localiza o arquivo EPS (aceitando .xlsx ou .csv)
    arquivo_eps = None
    extensao_eps = None
    
    if os.path.exists(pasta_eps):
        for arquivo in os.listdir(pasta_eps):
            if "EPS" in arquivo.upper() and nova_data in arquivo:
                if arquivo.endswith(".xlsx") or arquivo.endswith(".csv"):
                    arquivo_eps = os.path.join(pasta_eps, arquivo)
                    extensao_eps = os.path.splitext(arquivo)[1].lower()
                    break

    if not arquivo_eps:
        print(f"Erro: Arquivo EPS contendo 'EPS' e '{nova_data}' (.xlsx ou .csv) não foi encontrado em: {pasta_eps}")
        return

    print(f"Relatório EPS localizado: {os.path.basename(arquivo_eps)}")

    # 3. Carrega os dados do EPS usando Pandas mantendo formato textual para tratamento limpo
    try:
        if extensao_eps == ".csv":
            df_eps = pd.read_csv(arquivo_eps, sep=None, engine='python', encoding='latin-1', dtype=str)
        else:
            df_eps = pd.read_excel(arquivo_eps, dtype=str)
    except Exception as e:
        print(f"Erro ao ler o arquivo EPS: {e}")
        return

    if df_eps.empty:
        print("Aviso: O arquivo EPS localizado está vazio.")
        return

    # Mapeamento exato das colunas que precisam de tratamento numérico e de data
    colunas_numero = ["nVlrInicio", "nVlrResgate", "nDespApropriar", "nDespApropriada", "nDespDiaria",
                      "nVlrAtual", "nVlrContabilAtual", "nVlrComissaoCustoVendaApropriar", "nApropriacaoCustoVenda"]
    colunas_data = ["dtDataBase", "dtDtVcto", "dtDtEmissao"]

    indices_numero = [df_eps.columns.get_loc(col) for col in colunas_numero if col in df_eps.columns]
    indices_data = [df_eps.columns.get_loc(col) for col in colunas_data if col in df_eps.columns]

    cabecalho = list(df_eps.columns)
    linhas_dados = df_eps.values.tolist()

    print("Copiando dados do EPS com tratamento numérico real para a aba 'EMISSÃO PROPRIA SIMPLIFICADO'...")
    print("-" * 50)

    # 4. Varre todas as capas da pasta principal para colar o relatório
    contador_capas = 0
    for nome_arquivo in os.listdir(caminho_pasta_capas):
        caminho_capa = os.path.join(caminho_pasta_capas, nome_arquivo)

        if os.path.isfile(caminho_capa) and nome_arquivo.endswith(".xlsx"):
            try:
                wb = openpyxl.load_workbook(caminho_capa)

                if "EMISSÃO PROPRIA SIMPLIFICADO" in wb.sheetnames:
                    ws = wb["EMISSÃO PROPRIA SIMPLIFICADO"]

                    # Limpa completamente dados e formatações antigas a partir da linha 30
                    for linha_limpar in range(30, ws.max_row + 1):
                        for col_limpar in range(1, len(cabecalho) + 1):
                            ws.cell(row=linha_limpar, column=col_limpar).value = None
                            ws.cell(row=linha_limpar, column=col_limpar).number_format = 'General'

                    # 1. Cola o Cabeçalho na Linha 30
                    for c_idx, valor in enumerate(cabecalho):
                        celula = ws.cell(row=30, column=c_idx + 1)
                        celula.value = valor

                    # 2. Cola as Linhas de Dados a partir da Linha 31
                    for r_idx, linha_dados in enumerate(linhas_dados):
                        linha_atual = 31 + r_idx
                        for c_idx, valor in enumerate(linha_dados):
                            celula = ws.cell(row=linha_atual, column=c_idx + 1)
                            
                            if pd.isna(valor) or str(valor).strip() == "" or str(valor).lower() == "nan":
                                celula.value = None
                                continue

                            valor_str = str(valor).strip()

                            # --- TRATAMENTO NUMÉRICO CRÍTICO ---
                            if c_idx in indices_numero:
                                try:
                                    # Limpeza profunda de espaços e caracteres regionais
                                    valor_str = valor_str.replace(" ", "")
                                    
                                    # Trata se o sinal de menos estiver no fim (ex: 1.500,00-)
                                    if valor_str.endswith("-"):
                                        valor_str = "-" + valor_str[:-1]
                                        
                                    if "," in valor_str:
                                        valor_str = valor_str.replace(".", "").replace(",", ".")
                                    
                                    # CONVERSÃO PARA FLOAT REAL (Garante que vira número para fórmulas)
                                    celula.value = float(valor_str)
                                    celula.number_format = '#,##0.00'
                                except Exception:
                                    # Se falhar, tenta remover qualquer outro caractere residual
                                    try:
                                        limpo = re.sub(r'[^\d,.-]', '', valor_str)
                                        if "," in limpo:
                                            limpo = limpo.replace(".", "").replace(",", ".")
                                        celula.value = float(limpo)
                                        celula.number_format = '#,##0.00'
                                    except:
                                        celula.value = valor  # Fallback final
                                
                            # --- TRATAMENTO DE DATAS CRÍTICO ---
                            elif c_idx in indices_data:
                                try:
                                    data_objeto = pd.to_datetime(valor_str, dayfirst=True, errors='coerce')
                                    if pd.notna(data_objeto):
                                        # Injeta o objeto DATE nativo do Python
                                        celula.value = data_objeto.to_pydatetime().date()
                                        celula.number_format = 'yyyy-mm-dd' # Formato interno que o Excel traduz localmente
                                    else:
                                        celula.value = valor
                                except Exception:
                                    celula.value = valor
                                
                            # Demais colunas de texto normal
                            else:
                                celula.value = valor_str

                    # --- CORREÇÃO DO ### (AUTO-AJUSTE DA LARGURA DAS COLUNAS) ---
                    for col in ws.columns:
                        max_len = 0
                        col_letter = openpyxl.utils.get_column_letter(col[0].column)
                        for cell in col:
                            if cell.row >= 30 and cell.value:
                                max_len = max(max_len, len(str(cell.value)))
                        if max_len > 0:
                            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

                    wb.save(caminho_capa)
                    print(f"Dados do EPS formatados e colados em: '{nome_arquivo}'")
                    contador_capas += 1

                wb.close()
            except Exception as e:
                print(f"Erro ao colar EPS no arquivo '{nome_arquivo}': {e}")

    print("-" * 50)
    print(f"Processamento do EPS concluído! Total de {contador_capas} capas atualizadas.")

def importar_dados_opc(caminho_pasta_capas, nova_data):
    # 1. Define o caminho da pasta do relatório OPC (mesma pasta do EPC/EPS)
    pasta_opc = r"C:\Users\Hall\Documents\Python - HOMOLOGACAO\Capas Python\Arquivos"

    print(f"\n--- Buscando Relatório OPC para a data {nova_data} ---")

    # 2. Localiza o arquivo OPC (aceitando .xlsx ou .csv)
    arquivo_opc = None
    extensao_opc = None
    
    if os.path.exists(pasta_opc):
        for arquivo in os.listdir(pasta_opc):
            if "OPC" in arquivo.upper() and nova_data in arquivo:
                if arquivo.endswith(".xlsx") or arquivo.endswith(".csv"):
                    arquivo_opc = os.path.join(pasta_opc, arquivo)
                    extensao_opc = os.path.splitext(arquivo)[1].lower()
                    break

    if not arquivo_opc:
        print(f"Erro: Arquivo OPC contendo 'OPC' e '{nova_data}' (.xlsx ou .csv) não foi encontrado em: {pasta_opc}")
        return

    print(f"Relatório OPC localizado: {os.path.basename(arquivo_opc)}")

    # 3. Carrega os dados do OPC usando Pandas mantendo formato textual para tratamento limpo
    try:
        if extensao_opc == ".csv":
            df_opc = pd.read_csv(arquivo_opc, sep=None, engine='python', encoding='latin-1', dtype=str)
        else:
            df_opc = pd.read_excel(arquivo_opc, dtype=str)
    except Exception as e:
        print(f"Erro ao ler o arquivo OPC: {e}")
        return

    if df_opc.empty:
        print("Aviso: O arquivo OPC localizado está vazio.")
        return

    # Mapeamento exato das colunas que precisam de tratamento numérico e de data
    colunas_numero = ["nVlrAplicado", "nDespRendaApropriarTotal", "nDespRendaApropriar",
                      "nVlrAtual", "nVlrVolta", "nVlrContabil"]
    colunas_data = ["dtDtEmissao"]

    indices_numero = [df_opc.columns.get_loc(col) for col in colunas_numero if col in df_opc.columns]
    indices_data = [df_opc.columns.get_loc(col) for col in colunas_data if col in df_opc.columns]

    cabecalho = list(df_opc.columns)
    linhas_dados = df_opc.values.tolist()

    print("Copiando dados do OPC com tratamento numérico real para a aba 'OPERAÇÃO COMPROMISSADA'...")
    print("-" * 50)

    # 4. Varre todas as capas da pasta principal para colar o relatório
    contador_capas = 0
    for nome_arquivo in os.listdir(caminho_pasta_capas):
        caminho_capa = os.path.join(caminho_pasta_capas, nome_arquivo)

        if os.path.isfile(caminho_capa) and nome_arquivo.endswith(".xlsx"):
            try:
                wb = openpyxl.load_workbook(caminho_capa)

                if "OPERAÇÃO COMPROMISSADA" in wb.sheetnames:
                    ws = wb["OPERAÇÃO COMPROMISSADA"]

                    # Limpa completamente dados e formatações antigas a partir da linha 30
                    for linha_limpar in range(30, ws.max_row + 1):
                        for col_limpar in range(1, len(cabecalho) + 1):
                            ws.cell(row=linha_limpar, column=col_limpar).value = None
                            ws.cell(row=linha_limpar, column=col_limpar).number_format = 'General'

                    # 1. Cola o Cabeçalho na Linha 30
                    for c_idx, valor in enumerate(cabecalho):
                        celula = ws.cell(row=30, column=c_idx + 1)
                        celula.value = valor

                    # 2. Cola as Linhas de Dados a partir da Linha 31
                    for r_idx, linha_dados in enumerate(linhas_dados):
                        linha_atual = 31 + r_idx
                        for c_idx, valor in enumerate(linha_dados):
                            celula = ws.cell(row=linha_atual, column=c_idx + 1)
                            
                            if pd.isna(valor) or str(valor).strip() == "" or str(valor).lower() == "nan":
                                celula.value = None
                                continue

                            valor_str = str(valor).strip()

                            # --- TRATAMENTO NUMÉRICO CRÍTICO ---
                            if c_idx in indices_numero:
                                try:
                                    # Limpeza profunda de espaços e caracteres regionais
                                    valor_str = valor_str.replace(" ", "")
                                    
                                    # Trata se o sinal de menos estiver no fim (ex: 1.500,00-)
                                    if valor_str.endswith("-"):
                                        valor_str = "-" + valor_str[:-1]
                                        
                                    if "," in valor_str:
                                        valor_str = valor_str.replace(".", "").replace(",", ".")
                                    
                                    # CONVERSÃO PARA FLOAT REAL (Garante que vira número para fórmulas)
                                    celula.value = float(valor_str)
                                    celula.number_format = '#,##0.00'
                                except Exception:
                                    # Se falhar, tenta remover qualquer outro caractere residual
                                    try:
                                        limpo = re.sub(r'[^\d,.-]', '', valor_str)
                                        if "," in limpo:
                                            limpo = limpo.replace(".", "").replace(",", ".")
                                        celula.value = float(limpo)
                                        celula.number_format = '#,##0.00'
                                    except:
                                        celula.value = valor  # Fallback final
                                
                            # --- TRATAMENTO DE DATAS CRÍTICO ---
                            elif c_idx in indices_data:
                                try:
                                    data_objeto = pd.to_datetime(valor_str, dayfirst=True, errors='coerce')
                                    if pd.notna(data_objeto):
                                        # Injeta o objeto DATE nativo do Python
                                        celula.value = data_objeto.to_pydatetime().date()
                                        celula.number_format = 'yyyy-mm-dd' # Formato interno que o Excel traduz localmente
                                    else:
                                        celula.value = valor
                                except Exception:
                                    celula.value = valor
                                
                            # Demais colunas de texto normal
                            else:
                                celula.value = valor_str

                    # --- CORREÇÃO DO ### (AUTO-AJUSTE DA LARGURA DAS COLUNAS) ---
                    for col in ws.columns:
                        max_len = 0
                        col_letter = openpyxl.utils.get_column_letter(col[0].column)
                        for cell in col:
                            if cell.row >= 30 and cell.value:
                                max_len = max(max_len, len(str(cell.value)))
                        if max_len > 0:
                            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

                    wb.save(caminho_capa)
                    print(f"Dados do OPC formatados e colados em: '{nome_arquivo}'")
                    contador_capas += 1

                wb.close()
            except Exception as e:
                print(f"Erro ao colar OPC no arquivo '{nome_arquivo}': {e}")

    print("-" * 50)
    print(f"Processamento do OPC concluído! Total de {contador_capas} capas atualizadas.")

def importar_dados_est(caminho_pasta_capas, nova_data):
    # 1. Define o caminho da pasta do relatório EST (mesma pasta do EPC/EPS/OPC)
    pasta_est = r"C:\Users\Hall\Documents\Python - HOMOLOGACAO\Capas Python\Arquivos"

    print(f"\n--- Buscando Relatório EST para a data {nova_data} ---")

    # 2. Localiza o arquivo EST (aceitando .xlsx ou .csv)
    arquivo_est = None
    extensao_est = None
    
    if os.path.exists(pasta_est):
        for arquivo in os.listdir(pasta_est):
            if "EST" in arquivo.upper() and nova_data in arquivo:
                if arquivo.endswith(".xlsx") or arquivo.endswith(".csv"):
                    arquivo_est = os.path.join(pasta_est, arquivo)
                    extensao_est = os.path.splitext(arquivo)[1].lower()
                    break

    if not arquivo_est:
        print(f"Erro: Arquivo EST contendo 'EST' e '{nova_data}' (.xlsx ou .csv) não foi encontrado em: {pasta_est}")
        return

    print(f"Relatório EST localizado: {os.path.basename(arquivo_est)}")

    # 3. Carrega os dados do EST usando Pandas mantendo formato textual para tratamento limpo
    try:
        if extensao_est == ".csv":
            df_est = pd.read_csv(arquivo_est, sep=None, engine='python', encoding='latin-1', dtype=str)
        else:
            df_est = pd.read_excel(arquivo_est, dtype=str)
    except Exception as e:
        print(f"Erro ao ler o arquivo EST: {e}")
        return

    if df_est.empty:
        print("Aviso: O arquivo EST localizado está vazio.")
        return

    # Mapeamento exato das colunas que precisam de tratamento numérico e de data
    colunas_numero = ["QTD", "VLR_JUSTO", "VLR_MERCADO", "VLR_CUSTO", "VLR_CONTABIL",
                      "VLR_APROPRIACAO", "VLR_AJUSTE_MTM"]
    colunas_data = ["DT_VCTO", "DT_EMISSAO"]

    indices_numero = [df_est.columns.get_loc(col) for col in colunas_numero if col in df_est.columns]
    indices_data = [df_est.columns.get_loc(col) for col in colunas_data if col in df_est.columns]

    cabecalho = list(df_est.columns)
    linhas_dados = df_est.values.tolist()

    print("Copiando dados do EST com tratamento numérico real para a aba 'Posição de Estoque por Valor'...")
    print("-" * 50)

    # 4. Varre todas as capas da pasta principal para colar o relatório
    contador_capas = 0
    for nome_arquivo in os.listdir(caminho_pasta_capas):
        caminho_capa = os.path.join(caminho_pasta_capas, nome_arquivo)

        if os.path.isfile(caminho_capa) and nome_arquivo.endswith(".xlsx"):
            try:
                wb = openpyxl.load_workbook(caminho_capa)

                if "Posição de Estoque por Valor" in wb.sheetnames:
                    ws = wb["Posição de Estoque por Valor"]

                    # Limpa completamente dados e formatações antigas a partir da linha 30
                    for linha_limpar in range(30, ws.max_row + 1):
                        for col_limpar in range(1, len(cabecalho) + 1):
                            ws.cell(row=linha_limpar, column=col_limpar).value = None
                            ws.cell(row=linha_limpar, column=col_limpar).number_format = 'General'

                    # 1. Cola o Cabeçalho na Linha 30
                    for c_idx, valor in enumerate(cabecalho):
                        celula = ws.cell(row=30, column=c_idx + 1)
                        celula.value = valor

                    # 2. Cola as Linhas de Dados a partir da Linha 31
                    for r_idx, linha_dados in enumerate(linhas_dados):
                        linha_atual = 31 + r_idx
                        for c_idx, valor in enumerate(linha_dados):
                            celula = ws.cell(row=linha_atual, column=c_idx + 1)
                            
                            if pd.isna(valor) or str(valor).strip() == "" or str(valor).lower() == "nan":
                                celula.value = None
                                continue

                            valor_str = str(valor).strip()

                            # --- TRATAMENTO NUMÉRICO CRÍTICO ---
                            if c_idx in indices_numero:
                                try:
                                    # Limpeza profunda de espaços e caracteres regionais
                                    valor_str = valor_str.replace(" ", "")
                                    
                                    # Trata se o sinal de menos estiver no fim (ex: 1.500,00-)
                                    if valor_str.endswith("-"):
                                        valor_str = "-" + valor_str[:-1]
                                        
                                    if "," in valor_str:
                                        valor_str = valor_str.replace(".", "").replace(",", ".")
                                    
                                    # CONVERSÃO PARA FLOAT REAL (Garante que vira número para fórmulas)
                                    celula.value = float(valor_str)
                                    celula.number_format = '#,##0.00'
                                except Exception:
                                    # Se falhar, tenta remover qualquer outro caractere residual
                                    try:
                                        limpo = re.sub(r'[^\d,.-]', '', valor_str)
                                        if "," in limpo:
                                            limpo = limpo.replace(".", "").replace(",", ".")
                                        celula.value = float(limpo)
                                        celula.number_format = '#,##0.00'
                                    except:
                                        celula.value = valor  # Fallback final
                                
                            # --- TRATAMENTO DE DATAS CRÍTICO ---
                            elif c_idx in indices_data:
                                try:
                                    data_objeto = pd.to_datetime(valor_str, dayfirst=True, errors='coerce')
                                    if pd.notna(data_objeto):
                                        # Injeta o objeto DATE nativo do Python
                                        celula.value = data_objeto.to_pydatetime().date()
                                        celula.number_format = 'yyyy-mm-dd' # Formato interno que o Excel traduz localmente
                                    else:
                                        celula.value = valor
                                except Exception:
                                    celula.value = valor
                                
                            # Demais colunas de texto normal
                            else:
                                celula.value = valor_str

                    # --- CORREÇÃO DO ### (AUTO-AJUSTE DA LARGURA DAS COLUNAS) ---
                    for col in ws.columns:
                        max_len = 0
                        col_letter = openpyxl.utils.get_column_letter(col[0].column)
                        for cell in col:
                            if cell.row >= 30 and cell.value:
                                max_len = max(max_len, len(str(cell.value)))
                        if max_len > 0:
                            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

                    wb.save(caminho_capa)
                    print(f"Dados do EST formatados e colados em: '{nome_arquivo}'")
                    contador_capas += 1

                wb.close()
            except Exception as e:
                print(f"Erro ao colar EST no arquivo '{nome_arquivo}': {e}")

    print("-" * 50)
    print(f"Processamento do EST concluído! Total de {contador_capas} capas atualizadas.")

def importar_dados_venture_capital(caminho_pasta_capas):
    # 1. Define o caminho da pasta do relatório Venture Capital (mesma pasta dos demais relatórios)
    pasta_vc = r"C:\Users\Hall\Documents\Python - HOMOLOGACAO\Capas Python\Arquivos"

    print("\n--- Buscando Relatório Venture Capital ---")

    # 2. Localiza o arquivo Venture Capital (somente .xlsx, pois precisamos preservar fórmulas e
    #    formatação). Este relatório NÃO possui data no nome, diferente dos demais (EPC, EPS, OPC, EST).
    arquivo_vc = None

    if os.path.exists(pasta_vc):
        for arquivo in os.listdir(pasta_vc):
            if "VENTURE CAPITAL" in arquivo.upper() and arquivo.endswith(".xlsx"):
                arquivo_vc = os.path.join(pasta_vc, arquivo)
                break

    if not arquivo_vc:
        print(f"Erro: Arquivo Venture Capital contendo 'Venture Capital' (.xlsx) não foi encontrado em: {pasta_vc}")
        return

    print(f"Relatório Venture Capital localizado: {os.path.basename(arquivo_vc)}")

    # 3. Carrega o workbook de origem preservando fórmulas (data_only=False) e formatação
    try:
        wb_origem = openpyxl.load_workbook(arquivo_vc, data_only=False)
    except Exception as e:
        print(f"Erro ao ler o arquivo Venture Capital: {e}")
        return

    if "Venture Capital" not in wb_origem.sheetnames:
        print(f"Erro: A aba 'Venture Capital' não foi encontrada no arquivo de origem '{os.path.basename(arquivo_vc)}'.")
        return

    ws_origem = wb_origem["Venture Capital"]

    print("Copiando a planilha 'Venture Capital' (valores, fórmulas e formatação) a partir da célula A1...")
    print("-" * 50)

    # 4. Varre todas as capas da pasta principal para colar a planilha inteira
    contador_capas = 0
    for nome_arquivo in os.listdir(caminho_pasta_capas):
        caminho_capa = os.path.join(caminho_pasta_capas, nome_arquivo)

        if os.path.isfile(caminho_capa) and nome_arquivo.endswith(".xlsx"):
            try:
                wb = openpyxl.load_workbook(caminho_capa)

                if "Venture Capital" in wb.sheetnames:
                    ws = wb["Venture Capital"]

                    # Remove combinações de células (merged cells) antigas antes de limpar
                    for intervalo in list(ws.merged_cells.ranges):
                        ws.unmerge_cells(str(intervalo))

                    # Limpa completamente dados e formatações antigas a partir da célula A1
                    linhas_limpar = max(ws.max_row, ws_origem.max_row)
                    colunas_limpar = max(ws.max_column, ws_origem.max_column)
                    for linha_limpar in range(1, linhas_limpar + 1):
                        for col_limpar in range(1, colunas_limpar + 1):
                            celula_limpar = ws.cell(row=linha_limpar, column=col_limpar)
                            celula_limpar.value = None
                            celula_limpar.number_format = 'General'

                    # 1. Copia todas as células (valor/fórmula + formatação completa) a partir de A1
                    for linha in ws_origem.iter_rows():
                        for celula_origem in linha:
                            celula_destino = ws.cell(row=celula_origem.row, column=celula_origem.column)

                            # Cola o valor OU a fórmula (a fórmula é colada como texto e será
                            # recalculada automaticamente pelo Excel na planilha de destino,
                            # atualizando o resultado conforme os dados da conta contábil)
                            celula_destino.value = celula_origem.value

                            # Copia a formatação completa da célula de origem (fonte, borda,
                            # preenchimento, alinhamento, proteção e formato numérico original)
                            if celula_origem.has_style:
                                celula_destino.font = copy.copy(celula_origem.font)
                                celula_destino.border = copy.copy(celula_origem.border)
                                celula_destino.fill = copy.copy(celula_origem.fill)
                                celula_destino.alignment = copy.copy(celula_origem.alignment)
                                celula_destino.protection = copy.copy(celula_origem.protection)
                                celula_destino.number_format = celula_origem.number_format

                            # --- TRATAMENTO NUMÉRICO CRÍTICO (mesmo padrão do EPC) ---
                            # Aplica o formato de número do EPC em toda célula numérica ou
                            # fórmula (cujo resultado é numérico), mantendo o restante da
                            # formatação original (datas, texto, etc.) intacta.
                            valor = celula_origem.value
                            eh_formula = isinstance(valor, str) and valor.startswith("=")
                            eh_numero = isinstance(valor, (int, float)) and not isinstance(valor, bool)
                            if eh_numero or eh_formula:
                                celula_destino.number_format = '#,##0.00'

                    # 2. Copia as combinações de células (merged cells) da origem
                    for intervalo in ws_origem.merged_cells.ranges:
                        ws.merge_cells(str(intervalo))

                    # 3. Copia a largura das colunas
                    for col_letter, dimensao in ws_origem.column_dimensions.items():
                        ws.column_dimensions[col_letter].width = dimensao.width

                    # 4. Copia a altura das linhas
                    for indice_linha, dimensao in ws_origem.row_dimensions.items():
                        ws.row_dimensions[indice_linha].height = dimensao.height

                    wb.save(caminho_capa)
                    print(f"Planilha 'Venture Capital' copiada com sucesso em: '{nome_arquivo}'")
                    contador_capas += 1
                else:
                    print(f"Aviso: Aba 'Venture Capital' não encontrada em '{nome_arquivo}'.")

                wb.close()
            except Exception as e:
                print(f"Erro ao colar Venture Capital no arquivo '{nome_arquivo}': {e}")

    print("-" * 50)
    print(f"Processamento do Venture Capital concluído! Total de {contador_capas} capas atualizadas.")

def importar_dados_titulos_capitalizacao(caminho_pasta_capas):
    # 1. Define o caminho da pasta do relatório TITULOS DE CAPITALIZAÇÃO (mesma pasta dos demais relatórios)
    pasta_tc = r"C:\Users\Hall\Documents\Python - HOMOLOGACAO\Capas Python\Arquivos"

    print("\n--- Buscando Relatório TITULOS DE CAPITALIZAÇÃO ---")

    # 2. Localiza o arquivo (somente .xlsx, pois precisamos preservar fórmulas e formatação).
    #    Este relatório NÃO possui data no nome, diferente dos demais (EPC, EPS, OPC, EST).
    arquivo_tc = None

    if os.path.exists(pasta_tc):
        for arquivo in os.listdir(pasta_tc):
            if "TITULOS DE CAPITALIZA" in arquivo.upper() and arquivo.endswith(".xlsx"):
                arquivo_tc = os.path.join(pasta_tc, arquivo)
                break

    if not arquivo_tc:
        print(f"Erro: Arquivo TITULOS DE CAPITALIZAÇÃO (.xlsx) não foi encontrado em: {pasta_tc}")
        return

    print(f"Relatório TITULOS DE CAPITALIZAÇÃO localizado: {os.path.basename(arquivo_tc)}")

    # 3. Carrega o workbook de origem preservando fórmulas (data_only=False) e formatação
    try:
        wb_origem = openpyxl.load_workbook(arquivo_tc, data_only=False)
    except Exception as e:
        print(f"Erro ao ler o arquivo TITULOS DE CAPITALIZAÇÃO: {e}")
        return

    if "TITULOS DE CAPITALIZAÇÃO" not in wb_origem.sheetnames:
        print(f"Erro: A aba 'TITULOS DE CAPITALIZAÇÃO' não foi encontrada no arquivo de origem '{os.path.basename(arquivo_tc)}'.")
        return

    ws_origem = wb_origem["TITULOS DE CAPITALIZAÇÃO"]

    print("Copiando a planilha 'TITULOS DE CAPITALIZAÇÃO' (valores, fórmulas e formatação) a partir da célula A1...")
    print("-" * 50)

    # 4. Varre todas as capas da pasta principal para colar a planilha inteira
    contador_capas = 0
    for nome_arquivo in os.listdir(caminho_pasta_capas):
        caminho_capa = os.path.join(caminho_pasta_capas, nome_arquivo)

        if os.path.isfile(caminho_capa) and nome_arquivo.endswith(".xlsx"):
            try:
                wb = openpyxl.load_workbook(caminho_capa)

                if "TITULOS DE CAPITALIZAÇÃO" in wb.sheetnames:
                    ws = wb["TITULOS DE CAPITALIZAÇÃO"]

                    # Remove combinações de células (merged cells) antigas antes de limpar
                    for intervalo in list(ws.merged_cells.ranges):
                        ws.unmerge_cells(str(intervalo))

                    # Limpa completamente dados e formatações antigas a partir da célula A1
                    linhas_limpar = max(ws.max_row, ws_origem.max_row)
                    colunas_limpar = max(ws.max_column, ws_origem.max_column)
                    for linha_limpar in range(1, linhas_limpar + 1):
                        for col_limpar in range(1, colunas_limpar + 1):
                            celula_limpar = ws.cell(row=linha_limpar, column=col_limpar)
                            celula_limpar.value = None
                            celula_limpar.number_format = 'General'

                    # 1. Copia todas as células (valor/fórmula + formatação completa) a partir de A1
                    for linha in ws_origem.iter_rows():
                        for celula_origem in linha:
                            celula_destino = ws.cell(row=celula_origem.row, column=celula_origem.column)

                            # Cola o valor OU a fórmula (a fórmula é colada como texto e será
                            # recalculada automaticamente pelo Excel na planilha de destino,
                            # atualizando o resultado conforme os dados da conta contábil)
                            celula_destino.value = celula_origem.value

                            # Copia a formatação completa da célula de origem (fonte, borda,
                            # preenchimento, alinhamento, proteção e formato numérico original)
                            if celula_origem.has_style:
                                celula_destino.font = copy.copy(celula_origem.font)
                                celula_destino.border = copy.copy(celula_origem.border)
                                celula_destino.fill = copy.copy(celula_origem.fill)
                                celula_destino.alignment = copy.copy(celula_origem.alignment)
                                celula_destino.protection = copy.copy(celula_origem.protection)
                                celula_destino.number_format = celula_origem.number_format

                            # --- TRATAMENTO NUMÉRICO CRÍTICO (mesmo padrão do EPC) ---
                            # Aplica o formato de número do EPC em toda célula numérica ou
                            # fórmula (cujo resultado é numérico), mantendo o restante da
                            # formatação original (datas, texto, etc.) intacta.
                            valor = celula_origem.value
                            eh_formula = isinstance(valor, str) and valor.startswith("=")
                            eh_numero = isinstance(valor, (int, float)) and not isinstance(valor, bool)
                            if eh_numero or eh_formula:
                                celula_destino.number_format = '#,##0.00'

                    # 2. Copia as combinações de células (merged cells) da origem
                    for intervalo in ws_origem.merged_cells.ranges:
                        ws.merge_cells(str(intervalo))

                    # 3. Copia a largura das colunas
                    for col_letter, dimensao in ws_origem.column_dimensions.items():
                        ws.column_dimensions[col_letter].width = dimensao.width

                    # 4. Copia a altura das linhas
                    for indice_linha, dimensao in ws_origem.row_dimensions.items():
                        ws.row_dimensions[indice_linha].height = dimensao.height

                    wb.save(caminho_capa)
                    print(f"Planilha 'TITULOS DE CAPITALIZAÇÃO' copiada com sucesso em: '{nome_arquivo}'")
                    contador_capas += 1
                else:
                    print(f"Aviso: Aba 'TITULOS DE CAPITALIZAÇÃO' não encontrada em '{nome_arquivo}'.")

                wb.close()
            except Exception as e:
                print(f"Erro ao colar TITULOS DE CAPITALIZAÇÃO no arquivo '{nome_arquivo}': {e}")

    print("-" * 50)
    print(f"Processamento do TITULOS DE CAPITALIZAÇÃO concluído! Total de {contador_capas} capas atualizadas.")

if __name__ == "__main__":
    # 1. Executa a atualização de datas
    caminho, data_input = atualizar_datas_arquivos()

    from datetime import datetime
    data_formatada_bof = datetime.strptime(data_input, "%d/%m/%Y").strftime("%m-%Y")

    # 2. Preenche as informações fixas/assinaturas (CONCILIAÇÃO)
    preencher_informacoes_planilha(caminho, data_input)
    # Nota: quando executado via GUI (Dashboard_Geracao_de_Capas.py), os parâmetros
    # quem_concilia, quem_aprova e quem_reconcilia são passados diretamente e o pipeline
    # completo é orquestrado por lá, não por este bloco __main__.

    # 3. Importa e cruza os dados do Balancete (BOF)
    importar_dados_balancete(caminho, data_formatada_bof)
    
    # 4. Importa e replica o relatório completo EPC em todas as planilhas
    importar_dados_epc(caminho, data_formatada_bof)

    # 5. Importa e replica o relatório simplificado EPS em todas as planilhas
    importar_dados_eps(caminho, data_formatada_bof)

    # 6. Importa e replica o relatório de operação compromissada OPC em todas as planilhas
    importar_dados_opc(caminho, data_formatada_bof)

    # 7. Importa e replica o relatório de posição de estoque EST em todas as planilhas
    importar_dados_est(caminho, data_formatada_bof)

    # 8. Copia a planilha Venture Capital inteira (valores, fórmulas e formatação) em todas as capas
    importar_dados_venture_capital(caminho)

    # 9. Copia a planilha TITULOS DE CAPITALIZAÇÃO inteira (valores, fórmulas e formatação) em todas as capas
    importar_dados_titulos_capitalizacao(caminho)
  
    
  
    